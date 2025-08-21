import os
import pandas as pd
from openpyxl import load_workbook
from getFilename import find_latest_file


# -------------------------
# Extract hyperlinks from latest Excel file
# -------------------------
def extract_hyperlinks(folder_name: str, save_hyperlink: bool = True) -> pd.DataFrame | None:
    base_directory = os.path.dirname(__file__)
    folder_path = os.path.join(base_directory, '..', folder_name)

    latest_file = find_latest_file(folder_path)
    if not latest_file:
        print(f"No Excel file found in folder '{folder_name}'.")
        return None

    file_path = os.path.join(folder_path, latest_file)
    print(f"Processing latest file: {latest_file}")

    # --- Load workbook and sheet
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)

    # --- Locate MOA and ADDENDUM columns
    moa_idx = df.columns.get_loc("MOA")
    add_idx = df.columns.get_loc("ADDENDUM")

    extracted_moa, extracted_addendum = [], []

    # --- Extract hyperlinks row by row
    for row in ws.iter_rows(min_row=2):  # skip header
        cell_moa, cell_add = row[moa_idx], row[add_idx]
        extracted_moa.append(cell_moa.hyperlink.target if cell_moa.hyperlink else None)
        extracted_addendum.append(cell_add.hyperlink.target if cell_add.hyperlink else None)

    # --- Add extracted links to DataFrame
    df['Extracted Link'] = extracted_moa
    df['Extracted Addendum'] = extracted_addendum

    # --- Optionally save results
    if save_hyperlink:
        os.makedirs("Hyperlink", exist_ok=True)
        save_path = os.path.join("Hyperlink", f"hyperlink_{latest_file}")
        df.to_excel(save_path, index=False)
        print(f"Saved extracted hyperlinks to: {save_path}")

    return df
