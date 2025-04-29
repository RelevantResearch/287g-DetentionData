import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd

# Set your directory path
directory = r"after"

# This will collect all DataFrames
all_dataframes = []

for filename in os.listdir(directory):
    if filename.endswith(".xlsx") and filename.startswith("participatingAgencies"):
        filepath = os.path.join(directory, filename)
        wb = load_workbook(filepath)
        ws = wb.active

        # Detect the header row
        header = [cell.value for cell in ws[1]]
        if header is None:
            continue

        # Find MOA column index
        try:
            moa_idx = header.index("MOA") + 1  # openpyxl is 1-indexed
        except ValueError:
            continue  # Skip files without MOA column

        # Add 'Extracted Link' as a new column
        extracted_links = []
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=moa_idx)
            link = cell.hyperlink.target if cell.hyperlink else "-"
            extracted_links.append(link)

        # Read sheet into pandas and append the Extracted Link column
        df = pd.read_excel(filepath)
        df["Extracted Link"] = extracted_links[:len(df)]  # Make sure lengths match

        all_dataframes.append(df)

# Combine all DataFrames into one
if all_dataframes:
    final_df = pd.concat(all_dataframes, ignore_index=True)
    output_path = os.path.join(directory, "participatingAgencies_with_links.xlsx")
    final_df.to_excel(output_path, index=False)
    print(f"✅ Combined file saved as: {output_path}")
else:
    print("⚠️ No valid files found.")
