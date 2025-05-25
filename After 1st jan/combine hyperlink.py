import openpyxl

# Load the Excel workbook and select the active sheet
file_path = 'participatingAgencies-20250224183354.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Set headers for clarity
header_row = 1
link_column_name = 'MOA'
new_column_name = 'Extracted Link'

# Find the index of the 'MOA' column
moa_col_idx = None
for cell in ws[header_row]:
    if cell.value and cell.value.strip().lower() == link_column_name.lower():
        moa_col_idx = cell.column
        break

if moa_col_idx is None:
    raise ValueError(f"Column '{link_column_name}' not found.")

# Find the index for the new column
new_col_idx = ws.max_column + 1
ws.cell(row=header_row, column=new_col_idx).value = new_column_name

# Process rows to extract hyperlinks
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=moa_col_idx)
    if cell.hyperlink:
        ws.cell(row=row, column=new_col_idx).value = cell.hyperlink.target
    else:
        ws.cell(row=row, column=new_col_idx).value = "-"

# Save the updated file
new_file_path = 'participatingAgencies_with_links.xlsx'
wb.save(new_file_path)
print(f"Updated file saved as {new_file_path}")
